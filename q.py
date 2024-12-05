from pyrogram import Client, filters
import config, keyboards
import mysql.connector
from mysql.connector import Error
from datetime import datetime
import re
from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


bot = Client(
    api_id=config.API_ID,
    api_hash=config.API_HASH,
    bot_token=config.BOT_TOKEN,
    name="restaurant_bot"
)

try:
    connection = mysql.connector.connect(
        host=config.DB_HOST,
        user=config.DB_USER,
        password=config.DB_PASSWORD,
        database=config.DB_NAME
    )
    if connection.is_connected():
        print("Успішне підключення до бази даних")
except Error as e:
    print("Помилка при підключенні до MySQL", e)

user_states = {}


def button_filter(button):
    async def func(_, __, msg):
        return msg.text == button.text
    return filters.create(func, "ButtonFilter", button=button)


def state_filter(state):
    async def func(_, __, message):
        return user_states.get(message.from_user.id, {}).get('state') == state
    return filters.create(func, "StateFilter", state=state)


@bot.on_message(filters.command("start"))
async def start(bot, message):
    await message.reply('Привіт! Вибери, хто ти.', reply_markup=keyboards.kb_main)


@bot.on_message(button_filter(keyboards.btn_admin))
async def admin(bot, message):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM admins WHERE telegram_id = %s", (message.from_user.id,))
    user = cursor.fetchone()
    cursor.close()
    if user:
        await message.reply('Вiтаю', reply_markup=keyboards.kb_admin)
    else:
        await message.reply(f'Ви не адміністратор. Надайте ці цифри іншому адміністратору: {message.from_user.id}')


@bot.on_message(button_filter(keyboards.btn_restaurant))
async def restaurant(bot, message):
    user_id = message.from_user.id
    # Перевіряємо, чи є користувач рестораном
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM restaurants WHERE telegram_id = %s", (user_id,))
    restaurants = cursor.fetchone()
    cursor.close()
    print(restaurants)
    if restaurants:
        await message.reply("Вiтаю", reply_markup=keyboards.kb_restaurant)
    else:
        await message.reply(f'Ви не ресторан. Надайте ці цифри адміністратору: {message.from_user.id}')


@bot.on_message(button_filter(keyboards.btn_addAdmin))
async def addAdmin(bot, message):
    if message.from_user.id in config.ADMIN_IDS:
        await message.reply("Введіть дані адміністратора у форматі: id; username; ім'я; прізвище\n"
                            "Приклад: 630501685; @dzakrevsky; Данило; Закревський")
        user_states[message.from_user.id] = {'state': 'awaiting_admin_data'}
    else:
        await message.reply(f'Ви не адміністратор. Надайте ці цифри іншому адміністратору: {message.from_user.id}')


@bot.on_message(button_filter(keyboards.btn_addRestaurant))
async def addRestaurant(bot, message):
    # Перевіряємо, чи є користувач адміністратором
    cursor = connection.cursor()
    cursor.execute("SELECT 1 FROM admins WHERE telegram_id = %s", (message.from_user.id,))
    user = cursor.fetchone()
    cursor.close()
    if user:
        await message.reply(
            "Введіть дані ресторану у форматі: id; назва ресторану; адреса; телефон\n"
            "Приклад: 1; Ресторан А; вул. Хрещатик 1; +380123456789"
        )
        # Встановлюємо стан користувача
        user_states[message.from_user.id] = {'state': 'awaiting_restaurant_data'}
    else:
        await message.reply(
            f'Ви не адміністратор. Надайте ці цифри іншому адміністратору: {message.from_user.id}'
        )


@bot.on_message(button_filter(keyboards.btn_courier))
async def addCourier(bot, message):
    if message.from_user.id in config.ADMIN_IDS:
        await message.reply("Введіть дані кур'єра у форматі: id; username; ім'я; прізвище\n"
                            "Приклад: 123456789; @courier_username; Ім'я; Прізвище")
        user_states[message.from_user.id] = {'state': 'awaiting_courier_data'}
    else:
        await message.reply("У вас немає прав для додавання кур'єрів.")


@bot.on_message(filters.private & state_filter('awaiting_admin_data'))
async def receive_admin_data(bot, message):
    user_id = message.from_user.id
    # Отримуємо текст повідомлення
    text = message.text
    # Розбиваємо дані по роздільнику ';'
    data = text.split(';')
    if len(data) != 4:
        await message.reply("Неправильний формат даних. Будь ласка, використовуйте формат: id; username; ім'я; прізвище")
        return
    # Очищаємо пробіли
    data = [d.strip() for d in data]
    try:
        admin_id = int(data[0])
        username = data[1]
        first_name = data[2]
        last_name = data[3]
        # Вставляємо дані в базу даних
        cursor = connection.cursor()
        sql = """
            INSERT INTO admins (telegram_id, username, first_name, last_name)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                username = VALUES(username),
                first_name = VALUES(first_name),
                last_name = VALUES(last_name);
        """
        cursor.execute(sql, (admin_id, username, first_name, last_name))
        connection.commit()
        cursor.close()
        await message.reply("Адміністратора додано успішно!", reply_markup=keyboards.kb_admin)
        # Оновлюємо список адміністраторів у конфігурації
        if admin_id not in config.ADMIN_IDS:
            config.ADMIN_IDS.append(admin_id)
        # Скидаємо стан користувача
        user_states.pop(user_id, None)
    except ValueError:
        await message.reply("ID має бути числом. Будь ласка, перевірте формат даних.", reply_markup=keyboards.kb_admin)
    except mysql.connector.Error as err:
        await message.reply(f"Сталася помилка при додаванні адміністратора: {err}", reply_markup=keyboards.kb_admin)
    finally:
        # Скидаємо стан користувача, якщо сталася помилка
        user_states.pop(user_id, None)


@bot.on_message(filters.private & button_filter(keyboards.btn_checkOrder))
async def check_active_orders(bot, message):
    user_id = message.from_user.id
    if user_id not in config.ADMIN_IDS:
        await message.reply("У вас немає доступу до цієї функції.")
        return

    cursor = connection.cursor(dictionary=True)
    try:
        # Оновлений запит для отримання активних замовлень з полем courier_id
        cursor.execute("""
            SELECT o.id, o.client_name, o.phone_number, o.delivery_address, o.preparation_time, 
                   o.comment, o.is_paid, o.amount_due, o.courier_id, r.name AS restaurant_name
            FROM orders o
            JOIN restaurants r ON o.restaurant_id = r.id
            WHERE o.is_delivered = 0
            ORDER BY o.preparation_time DESC
        """)
        orders = cursor.fetchall()

        if not orders:
            await message.reply("Немає активних замовлень.")
            return

        # Надсилаємо кожне замовлення окремим повідомленням з inline-кнопками
        for order in orders:
            order_id = order['id']
            client_name = order['client_name']
            phone_number = order['phone_number']
            delivery_address = order['delivery_address']
            preparation_time = order['preparation_time'].strftime('%Y-%m-%d %H:%M')
            comment = order['comment']
            is_paid = order['is_paid']
            amount_due = order['amount_due']
            restaurant_name = order['restaurant_name']
            courier_id = order['courier_id']

            # Формуємо текст замовлення
            order_text = (
                f"🔹 **Замовлення #{order_id}**\n"
                f"🏢 Ресторан: {restaurant_name}\n"
                f"👤 Ім'я клієнта: {client_name}\n"
                f"📞 Телефон: {phone_number}\n"
                f"📍 Адреса доставки: {delivery_address}\n"
                f"⏰ Час приготування: {preparation_time}\n"
            )
            if comment:
                order_text += f"📝 Коментар: {comment}\n"
            order_text += (
                f"💰 Оплачено: {'Так' if is_paid else 'Ні'}\n"
            )
            if not is_paid and amount_due:
                order_text += f"💵 Сума до оплати: {amount_due}\n"
            order_text += f"📌 Статус доставки: **Не доставлено**\n\n"

            # Створюємо інлайн-клавіатуру з кнопками для замовлення
            btn_edit = InlineKeyboardButton('Редагувати замовлення', callback_data=f"edit_order_{order_id}")
            btn_send = InlineKeyboardButton('Відправити в чат', callback_data=f"send_order_{order_id}")

            if courier_id is None:
                # Якщо кур'єр не призначений, додаємо кнопку "Надіслати кур'єру"
                btn_send_worker = InlineKeyboardButton('Надіслати кур\'єру', callback_data=f"send_worker_order_{order_id}")
                markup = InlineKeyboardMarkup([
                    [btn_edit],
                    [btn_send],
                    [btn_send_worker]
                ])
            else:
                # Якщо кур'єр призначений, можна додати інші кнопки або залишити лише редагування та відправлення
                # Наприклад, додамо кнопку "Відмовитися від замовлення" або "Забрати замовлення"
                btn_take_order = InlineKeyboardButton("Забрати замовлення", callback_data=f"take_order_{order_id}")
                markup = InlineKeyboardMarkup([
                    [btn_take_order]
                ])

            # Надсилаємо повідомлення
            await bot.send_message(
                chat_id=user_id,
                text=order_text,
                reply_markup=markup
            )
    except mysql.connector.Error as err:
        await message.reply(f"Сталася помилка при отриманні замовлень: {err}")
    finally:
        cursor.close()


@bot.on_callback_query(filters.regex(r'^take_order_(\d+)$'))
async def take_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^take_order_(\d+)$', callback_query.data)
    print(1)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor(dictionary=True)
        try:
            # Отримуємо замовлення
            cursor.execute("""
                SELECT * FROM orders WHERE id = %s
            """, (order_id,))
            order = cursor.fetchone()

            if not order:
                await callback_query.answer("Замовлення не знайдено.", show_alert=True)
                return

            if not order['courier_id']:
                await callback_query.answer("Це замовлення вже не призначене кур'єру.", show_alert=True)
                return

            # Оновлюємо замовлення, видаляючи courier_id
            cursor.execute("""
                UPDATE orders
                SET courier_id = NULL
                WHERE id = %s
            """, (order_id,))
            connection.commit()

            await callback_query.answer("Замовлення успішно повернено.", show_alert=True)

            # Формуємо текст замовлення для адміністратора
            order_text = (
                f"🔄 **Замовлення #{order_id} повернено від кур'єра**\n"
                f"🏢 Ресторан: {order['restaurant_id']}\n"
                f"👤 Ім'я клієнта: {order['client_name']}\n"
                f"📞 Телефон: {order['phone_number']}\n"
                f"📍 Адреса доставки: {order['delivery_address']}\n"
                f"⏰ Час приготування: {order['preparation_time'].strftime('%Y-%m-%d %H:%M')}\n"
            )
            if order['comment']:
                order_text += f"📝 Коментар: {order['comment']}\n"
            order_text += (
                f"💰 Оплачено: {'Так' if order['is_paid'] else 'Ні'}\n"
            )
            if not order['is_paid'] and order['amount_due']:
                order_text += f"💵 Сума до оплати: {order['amount_due']}\n"
            order_text += f"📌 Статус доставки: **Не доставлено**\n\n"

            # Створюємо інлайн-клавіатуру з кнопками для замовлення
            btn_edit = InlineKeyboardButton('Редагувати замовлення', callback_data=f"edit_order_{order_id}")
            btn_send = InlineKeyboardButton('Відправити в чат', callback_data=f"send_order_{order_id}")
            btn_send_worker = InlineKeyboardButton('Надіслати кур\'єру', callback_data=f"send_worker_order_{order_id}")
            markup = InlineKeyboardMarkup([
                [btn_edit],
                [btn_send],
                [btn_send_worker]
            ])

            # Відправляємо замовлення адміністратору
            await bot.send_message(
                chat_id=user_id,
                text=order_text,
                reply_markup=markup
            )

        except mysql.connector.Error as err:
            print(f"Помилка в take_order_callback: {err}")
            await callback_query.answer("Сталася помилка при поверненні замовлення.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)


@bot.on_message(filters.private & state_filter('awaiting_restaurant_data'))
async def receive_restaurant_data(bot, message):
    user_id = message.from_user.id
    text = message.text
    data = text.split(';')
    if len(data) != 4:
        await message.reply(
            "Неправильний формат даних. Будь ласка, використовуйте формат: id; назва ресторану; адреса; телефон"
        )
        return
    # Очищаємо пробіли
    data = [d.strip() for d in data]
    try:
        restaurant_id = int(data[0])
        name = data[1]
        address = data[2]
        phone = data[3]
        telegram_id = message.from_user.id  # Отримуємо telegram_id ресторану
        # Вставляємо дані в базу даних
        cursor = connection.cursor()
        sql = """
            INSERT INTO restaurants (id, name, address, phone, telegram_id)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                name = VALUES(name),
                address = VALUES(address),
                phone = VALUES(phone),
                telegram_id = VALUES(telegram_id);
        """
        cursor.execute(sql, (restaurant_id, name, address, phone, telegram_id))
        connection.commit()
        cursor.close()
        await message.reply("Ресторан додано успішно!", reply_markup=keyboards.kb_admin)
    except ValueError:
        await message.reply(
            "ID має бути числом. Будь ласка, перевірте формат даних.",
            reply_markup=keyboards.kb_admin
        )
    except mysql.connector.Error as err:
        await message.reply(
            f"Сталася помилка при додаванні ресторану: {err}",
            reply_markup=keyboards.kb_admin
        )
    finally:
        # Скидаємо стан користувача
        user_states.pop(user_id, None)

@bot.on_message(filters.private & state_filter('awaiting_courier_data'))
async def receive_courier_data(bot, message):
    user_id = message.from_user.id
    text = message.text
    data = text.split(';')
    if len(data) != 4:
        await message.reply("Неправильний формат даних. Будь ласка, використовуйте формат: id; username; ім'я; прізвище")
        return
    data = [d.strip() for d in data]
    try:
        courier_id = int(data[0])
        username = data[1]
        first_name = data[2]
        last_name = data[3]
        # Вставляємо дані в базу даних
        cursor = connection.cursor()
        sql = """
            INSERT INTO couriers (telegram_id, username, first_name, last_name)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                username = VALUES(username),
                first_name = VALUES(first_name),
                last_name = VALUES(last_name);
        """
        cursor.execute(sql, (courier_id, username, first_name, last_name))
        connection.commit()
        cursor.close()
        await message.reply("Кур'єра додано успішно!", reply_markup=keyboards.kb_admin)
    except ValueError:
        await message.reply("ID має бути числом. Будь ласка, перевірте формат даних.")
    except mysql.connector.Error as err:
        await message.reply(f"Сталася помилка при додаванні кур'єра: {err}")
    finally:
        # Скидаємо стан користувача
        user_states.pop(user_id, None)

@bot.on_message(button_filter(keyboards.btn_showOrders))
async def show_orders(bot, message):
    user_id = message.from_user.id
    # Перевіряємо, чи є користувач рестораном
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM restaurants WHERE telegram_id = %s", (user_id,))
    restaurant = cursor.fetchone()
    if restaurant:
        restaurant_id = restaurant[0]
        # Отримуємо замовлення, які не відзначені як приготовані
        cursor.execute("""
            SELECT id, client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due
            FROM orders
            WHERE restaurant_id = %s AND is_prepared = 0
        """, (restaurant_id,))
        orders = cursor.fetchall()
        if orders:
            for order in orders:
                order_id = order[0]
                client_name = order[1]
                phone_number = order[2]
                delivery_address = order[3]
                preparation_time = order[4].strftime('%Y-%m-%d %H:%M')
                comment = order[5]
                is_paid = order[6]
                amount_due = order[7]
                # Формуємо текст повідомлення
                order_text = f"Замовлення #{order_id}\n"
                order_text += f"Ім'я клієнта: {client_name}\n"
                order_text += f"Телефон: {phone_number}\n"
                order_text += f"Адреса доставки: {delivery_address}\n"
                order_text += f"Час приготування: {preparation_time}\n"
                if comment:
                    order_text += f"Коментар: {comment}\n"
                order_text += f"Оплачено: {'Так' if is_paid else 'Ні'}\n"
                if not is_paid and amount_due:
                    order_text += f"Сума до оплати: {amount_due}\n"
                # Створюємо інлайн-клавіатуру з кнопками для замовлення
                btn_edit = InlineKeyboardButton('Редагувати замовлення', callback_data=f"edit_order_{order_id}")
                btn_send = InlineKeyboardButton('Відправити в чат', callback_data=f"send_order_{order_id}")
                btn_send_worker = InlineKeyboardButton('Надіслати курєру', callback_data=f"send_worker_order_{order_id}")
                markup = InlineKeyboardMarkup([
                    [btn_edit],
                    [btn_send],
                    [btn_send_worker]
                ])
                await bot.send_message(chat_id=user_id, text=order_text, reply_markup=markup)
        else:
            await message.reply("Немає замовлень, що очікують приготування.")
    else:
        await message.reply("У вас немає доступу до цієї функції.")
    cursor.close()

@bot.on_message(button_filter(keyboards.btn_view_courier_orders))
async def view_courier_orders(bot, message):
    user_id = message.from_user.id
    cursor = connection.cursor()
    try:
        # Перевіряємо, чи є користувач кур'єром
        cursor.execute("SELECT id FROM couriers WHERE telegram_id = %s", (user_id,))
        courier = cursor.fetchone()
        if courier:
            courier_id = courier[0]
            # Отримуємо замовлення, які призначені цьому кур'єру та ще не доставлені
            cursor.execute("""
                SELECT id, client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due
                FROM orders
                WHERE courier_id = %s AND (is_delivered = 0 OR is_delivered IS NULL)
            """, (courier_id,))
            orders = cursor.fetchall()
            if orders:
                for order in orders:
                    order_id = order[0]
                    client_name = order[1]
                    phone_number = order[2]
                    delivery_address = order[3]
                    preparation_time = order[4].strftime('%Y-%m-%d %H:%M')
                    comment = order[5]
                    is_paid = order[6]
                    amount_due = order[7]
                    # Формуємо текст повідомлення
                    order_text = f"Замовлення #{order_id}\n"
                    order_text += f"Ім'я клієнта: {client_name}\n"
                    order_text += f"Телефон: {phone_number}\n"
                    order_text += f"Адреса доставки: {delivery_address}\n"
                    order_text += f"Час приготування: {preparation_time}\n"
                    if comment:
                        order_text += f"Коментар: {comment}\n"
                    order_text += f"Оплачено: {'Так' if is_paid else 'Ні'}\n"
                    if not is_paid and amount_due:
                        order_text += f"Сума до оплати: {amount_due}\n"
                    # Додаємо кнопку для відмітки замовлення як доставленого
                    markup = InlineKeyboardMarkup(
                        [[InlineKeyboardButton("Замовлення доставлено", callback_data=f"order_delivered_{order_id}")]]
                    )
                    await bot.send_message(chat_id=user_id, text=order_text, reply_markup=markup)
            else:
                await message.reply("У вас немає активних замовлень.")
        else:
            await message.reply("У вас немає доступу до цієї функції.")
    except Exception as e:
        print(f"Помилка в view_courier_orders: {e}")
        await message.reply("Сталася помилка при отриманні замовлень.")
    finally:
        cursor.close()

@bot.on_callback_query(filters.regex(r'^order_delivered_(\d+)$'))
async def order_delivered_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^order_delivered_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        try:
            courier_id = callback_query.from_user.id
            # Перевіряємо, чи призначене замовлення цьому кур'єру
            cursor.execute("SELECT courier_id FROM orders WHERE id = %s", (order_id,))
            order = cursor.fetchone()
            if order:
                assigned_courier_id = order[0]
                if assigned_courier_id == courier_id:
                    # Відзначаємо замовлення як доставлене
                    cursor.execute("""
                        UPDATE orders
                        SET is_delivered = TRUE
                        WHERE id = %s
                    """, (order_id,))
                    connection.commit()
                    await callback_query.answer("Замовлення відзначено як доставлене.", show_alert=True)
                    # Видаляємо кнопку з повідомлення
                    await callback_query.message.edit_reply_markup(reply_markup=None)
                else:
                    await callback_query.answer("Це замовлення вам не призначено.", show_alert=True)
            else:
                await callback_query.answer("Замовлення не знайдено.", show_alert=True)
        except Exception as e:
            print(f"Помилка в order_delivered_callback: {e}")
            await callback_query.answer("Сталася помилка при оновленні замовлення.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)

@bot.on_callback_query(filters.regex(r'^accept_order_(\d+)$'))
async def accept_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^accept_order_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        try:
            courier_telegram_id = user_id
            # Перевіряємо, чи призначене замовлення іншому кур'єру
            cursor.execute("SELECT courier_id FROM orders WHERE id = %s", (order_id,))
            order = cursor.fetchone()
            if order:
                existing_courier_id = order[0]
                if existing_courier_id is None:
                    # Призначаємо замовлення кур'єру
                    # Спочатку знайдемо courier_id у таблиці couriers за telegram_id
                    cursor.execute("SELECT id FROM couriers WHERE telegram_id = %s", (courier_telegram_id,))
                    courier = cursor.fetchone()
                    if courier:
                        courier_id_db = courier[0]
                        cursor.execute("""
                            UPDATE orders
                            SET courier_id = %s
                            WHERE id = %s
                        """, (courier_id_db, order_id))
                        connection.commit()

                        # Отримуємо деталі замовлення для відправки кур'єру
                        cursor.execute("""
                            SELECT client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due, restaurant_id
                            FROM orders
                            WHERE id = %s
                        """, (order_id,))
                        order_details = cursor.fetchone()
                        if order_details:
                            client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due, restaurant_id = order_details
                            # Отримуємо інформацію про ресторан
                            cursor.execute("SELECT name, address, phone FROM restaurants WHERE id = %s", (restaurant_id,))
                            restaurant_info = cursor.fetchone()
                            if restaurant_info:
                                restaurant_name, restaurant_address, restaurant_phone = restaurant_info
                            else:
                                restaurant_name = restaurant_address = restaurant_phone = "Невідома"

                            # Форматуємо повідомлення для кур'єра
                            order_message = f"""
Ви прийняли замовлення #{order_id} від ресторану {restaurant_name}.

Деталі замовлення:
Ім'я клієнта: {client_name}
Номер телефону: {phone_number}
Адреса доставки: {delivery_address}
Час приготування замовлення: {preparation_time}
Коментар: {comment if comment else 'Немає'}

Чи сплачено замовлення: {'Так' if is_paid else 'Ні'}
Сума до сплати: {amount_due if not is_paid else 'Сплачено'}
                            """
                            # Відправляємо повідомлення кур'єру
                            await bot.send_message(chat_id=courier_telegram_id, text=order_message)

                            # Повідомляємо в чат, що замовлення прийняте
                            await callback_query.answer("Ви прийняли замовлення.", show_alert=True)
                            await callback_query.message.edit_reply_markup(reply_markup=None)
                        else:
                            await callback_query.answer("Деталі замовлення не знайдено.", show_alert=True)
                    else:
                        await callback_query.answer("Ви не зареєстровані як кур'єр.", show_alert=True)
                else:
                    await callback_query.answer("Замовлення вже прийняте іншим кур'єром.", show_alert=True)
            else:
                await callback_query.answer("Замовлення не знайдено.", show_alert=True)
        except Exception as e:
            print(f"Помилка в accept_order_callback: {e}")
            await callback_query.answer("Сталася помилка при прийнятті замовлення.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)

@bot.on_callback_query(filters.regex(r'^send_order_(\d+)$'))
async def send_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^send_order_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        try:
            # Перевіряємо, чи є користувач адміністратором
            cursor.execute("SELECT id FROM admins WHERE telegram_id = %s", (user_id,))
            admin = cursor.fetchone()
            if admin:
                # Отримуємо деталі замовлення
                cursor.execute("""
                    SELECT o.client_name, o.phone_number, o.delivery_address, o.preparation_time, o.comment, o.is_paid, 
                           o.amount_due,
                           r.name AS restaurant_name, r.address AS restaurant_address, r.phone AS restaurant_phone 
                    FROM orders o
                    JOIN restaurants r ON o.restaurant_id = r.id
                    WHERE o.id = %s
                """, (order_id,))
                order = cursor.fetchone()
                if order:
                    (client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due,
                     restaurant_name, restaurant_address, restaurant_phone) = order
                    # Форматуємо повідомлення
                    order_message = f"""
Новe замовлення від ресторану {restaurant_name}
Адреса ресторану: {restaurant_address}
Контактний номер ресторану: {restaurant_phone}

Ім'я клієнта: {client_name}
Номер телефону: {phone_number}
Адреса доставки: {delivery_address}
Час приготування замовлення: {preparation_time}
Коментар: {comment if comment else 'Немає'}

Чи сплачено замовлення: {'Так' if is_paid else 'Ні'}
Сума до сплати: {amount_due if not is_paid else 'Сплачено'}
                    """
                    # Створюємо інлайн-кнопку «Прийняти замовлення»
                    accept_button = InlineKeyboardButton('Прийняти замовлення', callback_data=f"accept_order_{order_id}")
                    inline_kb = InlineKeyboardMarkup([[accept_button]])

                    # Відправляємо повідомлення в вказаний чат
                    chat_id = -4578652782  # Замість цього використовуйте правильний ID чату (ціле число, без лапок)
                    await bot.send_message(chat_id, order_message, reply_markup=inline_kb)

                    await callback_query.answer("Замовлення надіслано в чат.", show_alert=True)
                else:
                    await callback_query.answer("Замовлення не знайдено.", show_alert=True)
            else:
                await callback_query.answer("У вас немає доступу до цієї функції.", show_alert=True)
        except Exception as e:
            print(f"Помилка в send_order_callback: {e}")
            await callback_query.answer("Сталася помилка при надсиланні замовлення в чат.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)


@bot.on_callback_query(filters.regex(r'^order_prepared_(\d+)$'))
async def order_prepared_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^order_prepared_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        # Перевіряємо, чи є користувач рестораном
        cursor = connection.cursor()
        cursor.execute("SELECT id FROM restaurants WHERE telegram_id = %s", (user_id,))
        restaurant = cursor.fetchone()
        if restaurant:
            restaurant_id = restaurant[0]
            # Оновлюємо статус замовлення на "приготоване"
            cursor.execute("""
                UPDATE orders
                SET is_prepared = TRUE
                WHERE id = %s AND restaurant_id = %s
            """, (order_id, restaurant_id))
            connection.commit()
            cursor.close()
            await callback_query.answer("Замовлення відзначено як приготоване.", show_alert=True)
            # Видаляємо кнопку з повідомлення
            await callback_query.message.edit_reply_markup(reply_markup=None)
        else:
            await callback_query.answer("У вас немає доступу до цієї функції.", show_alert=True)
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)


@bot.on_callback_query(filters.regex(r'^send_worker_order_(\d+)$'))
async def send_worker_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^send_worker_order_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        try:
            # Перевіряємо, чи є користувач адміністратором
            cursor.execute("SELECT id FROM admins WHERE telegram_id = %s", (user_id,))
            admin = cursor.fetchone()
            if admin:
                # Перевіряємо, чи існує замовлення
                cursor.execute("SELECT id FROM orders WHERE id = %s", (order_id,))
                order = cursor.fetchone()
                if order:
                    # Запитуємо Telegram ID кур'єра
                    await bot.send_message(
                        chat_id=user_id,
                        text=f"Введіть Telegram ID кур'єра для замовлення #{order_id}:"
                    )
                    # Встановлюємо стан очікування вводу Telegram ID кур'єра
                    if user_id not in user_states:
                        user_states[user_id] = {}
                    user_states[user_id]['state'] = 'awaiting_courier_id_assignment'
                    user_states[user_id]['order_id'] = order_id
                    await callback_query.answer("Введіть Telegram ID кур'єра.", show_alert=True)
                else:
                    await callback_query.answer("Замовлення не знайдено.", show_alert=True)
            else:
                await callback_query.answer("У вас немає доступу до цієї функції.", show_alert=True)
        except Exception as e:
            print(f"Помилка в send_worker_order_callback: {e}")
            await callback_query.answer("Сталася помилка.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)


@bot.on_message(filters.private & button_filter(keyboards.btn_archive))
async def check_archive_orders(bot, message):
    user_id = message.from_user.id
    # Перевіряємо, чи є користувач адміністратором
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT 1 FROM admins WHERE telegram_id = %s", (user_id,))
        admin = cursor.fetchone()
        if not admin:
            await message.reply("У вас немає доступу до цієї функції.")
            return

        # Отримуємо завершені замовлення (наприклад, is_delivered = 1)
        cursor.execute("""
                SELECT 
                    o.order_date, 
                    r.name AS restaurant_name, 
                    r.address AS restaurant_address,
                    o.cashless AS cashless, 
                    o.cash AS cash
                FROM orders o
                JOIN restaurants r ON o.restaurant_id = r.id
                WHERE o.is_delivered = 1
                ORDER BY o.order_date DESC
            """)
        orders = cursor.fetchall()

        if not orders:
            await message.reply("Немає завершених замовлень.")
            return

        # Створюємо новий робочий лист Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Архів Замовлень"

        # Визначаємо заголовки стовпців
        headers = ["Дата замовлення", "Заклад", "Адреса", "Безготівка", "Готівка"]
        ws.append(headers)

        # Заповнюємо дані замовлень
        for order in orders:
            order_date = order['order_date'].strftime('%d/%m/%Y') if order['order_date'] else ""
            restaurant = order['restaurant_name'] if order['restaurant_name'] else ""
            address = order['restaurant_address'] if order['restaurant_address'] else ""
            cashless = order['cashless'] if order['cashless'] is not None else 0
            cash = order['cash'] if order['cash'] is not None else 0
            ws.append([order_date, restaurant, address, cashless, cash])

        # Автоматичне налаштування ширини стовпців
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = length + 2  # Додаємо трохи простору

        # Зберігаємо Excel файл в пам'яті
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)  # Переміщаємо курсор на початок файлу

        # Відправляємо файл у чат
        await bot.send_document(
            chat_id=user_id,
            document=excel_file,
            file_name=f"archive_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            caption="Архів завершених замовлень."
        )

    except mysql.connector.Error as err:
        await message.reply(f"Сталася помилка при отриманні архіву: {err}")
    except Exception as e:
        await message.reply(f"Сталася непередбачена помилка: {e}")
    finally:
        cursor.close()


@bot.on_message(button_filter(keyboards.btn_newOrder))
async def createOrder(bot, message):
    if message.from_user.id in config.ADMIN_IDS:
        user_states[message.from_user.id] = {'state': 'awaiting_client_name', 'order_data': {}}
        await message.reply("Введіть ім'я клієнта:")
    else:
        await message.reply("У вас немає прав для створення замовлень.")


@bot.on_message(filters.private & state_filter('awaiting_client_name'))
async def get_client_name(bot, message):
    user_id = message.from_user.id
    user_states[user_id]['order_data']['client_name'] = message.text.strip()
    user_states[user_id]['state'] = 'awaiting_phone_number'
    await message.reply("Введіть номер телефону клієнта:")


@bot.on_message(filters.private & state_filter('awaiting_phone_number'))
async def get_phone_number(bot, message):
    user_id = message.from_user.id
    user_states[user_id]['order_data']['phone_number'] = message.text.strip()
    user_states[user_id]['state'] = 'awaiting_delivery_address'
    await message.reply("Введіть адресу доставки:")


@bot.on_message(filters.private & state_filter('awaiting_delivery_address'))
async def get_delivery_address(bot, message):
    user_id = message.from_user.id
    user_states[user_id]['order_data']['delivery_address'] = message.text.strip()
    user_states[user_id]['state'] = 'awaiting_preparation_time'
    await message.reply("Введіть час приготування замовлення (у форматі YYYY-MM-DD HH:MM):")


@bot.on_message(filters.private & state_filter('awaiting_preparation_time'))
async def get_preparation_time(bot, message):
    user_id = message.from_user.id
    try:
        prep_time = datetime.strptime(message.text.strip(), '%Y-%m-%d %H:%M')
        user_states[user_id]['order_data']['preparation_time'] = prep_time.strftime('%Y-%m-%d %H:%M:%S')
        user_states[user_id]['state'] = 'awaiting_comment'
        await message.reply("Введіть коментар до замовлення (необов'язково). Якщо немає, введіть 'немає':")
    except ValueError:
        await message.reply("Неправильний формат дати. Будь ласка, введіть час у форматі YYYY-MM-DD HH:MM")


@bot.on_message(filters.private & state_filter('awaiting_comment'))
async def get_comment(bot, message):
    user_id = message.from_user.id
    comment = message.text.strip()
    if comment.lower() == 'немає':
        comment = ''
    user_states[user_id]['order_data']['comment'] = comment
    user_states[user_id]['state'] = 'awaiting_is_paid'
    await message.reply("Чи оплачене замовлення? Введіть 'так' або 'ні':")


@bot.on_message(filters.private & state_filter('awaiting_is_paid'))
async def get_is_paid(bot, message):
    user_id = message.from_user.id
    response = message.text.strip().lower()
    if response == 'так':
        user_states[user_id]['order_data']['is_paid'] = True
        user_states[user_id]['order_data']['amount_due'] = None
        # Сохраняем заказ в базе данных
        await save_order(bot, message)
    elif response == 'ні':
        user_states[user_id]['order_data']['is_paid'] = False
        user_states[user_id]['state'] = 'awaiting_amount_due'
        await message.reply("Скільки потрібно взяти з клієнта?")
    else:
        await message.reply("Будь ласка, введіть 'так' або 'ні':")


@bot.on_message(filters.private & state_filter('awaiting_amount_due'))
async def get_amount_due(bot, message):
    user_id = message.from_user.id
    try:
        amount_due = float(message.text.strip().replace(',', '.'))
        user_states[user_id]['order_data']['amount_due'] = amount_due
        # Сохраняем заказ в базе данных
        await save_order(bot, message)
    except ValueError:
        await message.reply("Будь ласка, введіть числове значення суми.")


async def save_order(bot, message):
    user_id = message.from_user.id
    order_data = user_states[user_id]['order_data']

    # Automatically assign restaurant_id as the user ID (who sent the message)
    restaurant_id = user_id

    try:
        cursor = connection.cursor()
        sql = """
            INSERT INTO orders (client_name, phone_number, delivery_address, preparation_time, comment, is_paid, 
            amount_due, restaurant_id, order_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (
            order_data['client_name'],
            order_data['phone_number'],
            order_data['delivery_address'],
            order_data['preparation_time'],
            order_data['comment'],
            order_data['is_paid'],
            order_data['amount_due'],
            order_data['order_date'],
            restaurant_id  # Automatically assign restaurant_id as the user's ID
        ))
        connection.commit()
        cursor.close()
        await message.reply("Замовлення успішно створено!", reply_markup=keyboards.kb_restaurant)
        await new_order_alarm(bot, message)
    except mysql.connector.Error as err:
        await message.reply(f"Сталася помилка при збереженні замовлення: {err}")
    finally:
        # Clear the user's state after the order is saved
        user_states.pop(user_id, None)


@bot.on_callback_query(filters.regex(r'^edit_order_(\d+)$'))
async def edit_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^edit_order_(\d+)$', callback_query.data)

    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        cursor.execute("SELECT id FROM admins WHERE telegram_id = %s", (user_id,))
        restaurant = cursor.fetchone()
        if restaurant:
            # Запрашиваем текущие данные о заказе
            cursor.execute(
                "SELECT client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due FROM orders WHERE id = %s",
                (order_id,))
            order = cursor.fetchone()

            if order:
                client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due = order

                # Отправляем пользователю запрос на редактирование
                edit_msg = f"""
Редагування замовлення (ID: {order_id}):
Ім'я клієнта: {client_name}
Номер телефону: {phone_number}
Адреса доставки: {delivery_address}
Час приготування: {preparation_time}
Коментар: {comment}
Чи сплачено: {'Так' if is_paid else 'Ні'}
Сума до сплати: {amount_due if not is_paid else 'Сплачено'}
                """
                await bot.send_message(user_id, edit_msg)
                await callback_query.answer("Відправте нові дані для редагування замовлення."
                                            "Щоб змінити будь-яке з цих полів, відправте нові дані у форматі:"
                                            "Ім'я клієнта: [нове ім'я]", show_alert=True)

                if user_id not in user_states:
                    user_states[user_id] = {}

                # Ожидаем дальнейших данных для редактирования
                user_states[user_id]['editing_order_id'] = order_id  # Сохраняем, что пользователь редактирует заказ
            else:
                await callback_query.answer("Замовлення не знайдено.", show_alert=True)
        else:
            await callback_query.answer("У вас немає доступу до цієї функції.", show_alert=True)
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)


@bot.on_message(filters.private & state_filter('awaiting_courier_id_assignment'))
async def assign_order_to_courier(bot, message):
    user_id = message.from_user.id
    if user_id in user_states and user_states[user_id].get('state') == 'awaiting_courier_id_assignment':
        order_id = user_states[user_id].get('order_id')
        courier_telegram_id_input = message.text.strip()
        try:
            # Перевіряємо, чи є введений ID числом
            courier_telegram_id = int(courier_telegram_id_input)
        except ValueError:
            await message.reply("Неправильний формат ID. Будь ласка, введіть числовий Telegram ID кур'єра.")
            return

        cursor = connection.cursor()
        try:
            # Перевіряємо, чи користувач з таким Telegram ID існує
            # Ми не використовуємо таблицю couriers, тому просто перевіримо, чи може бот надсилати повідомлення цьому ID
            # Наприклад, спробуємо отримати користувача
            try:
                user = await bot.get_users(courier_telegram_id)
            except Exception as e:
                print(f"Помилка при отриманні користувача з ID {courier_telegram_id}: {e}")
                await message.reply("Не вдалося знайти користувача з таким Telegram ID.")
                return

            # Призначаємо замовлення кур'єру, зберігаючи Telegram ID
            cursor.execute("""
                UPDATE orders
                SET courier_id = %s
                WHERE id = %s
            """, (courier_telegram_id, order_id))
            connection.commit()

            # Отримуємо деталі замовлення
            cursor.execute("""
                SELECT client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due, restaurant_id
                FROM orders
                WHERE id = %s
            """, (order_id,))
            order_details = cursor.fetchone()
            if order_details:
                client_name, phone_number, delivery_address, preparation_time, comment, is_paid, amount_due, restaurant_id = order_details
                # Отримуємо інформацію про ресторан
                cursor.execute("SELECT name, address, phone FROM restaurants WHERE id = %s", (restaurant_id,))
                restaurant_info = cursor.fetchone()
                if restaurant_info:
                    restaurant_name, restaurant_address, restaurant_phone = restaurant_info
                else:
                    restaurant_name = restaurant_address = restaurant_phone = "Невідома"

                # Форматуємо повідомлення для кур'єра
                order_message = f"""
Ви були призначені на доставку замовлення #{order_id} від ресторану {restaurant_name}.
Адреса ресторану: {restaurant_address}
Контактний номер ресторану: {restaurant_phone}

Деталі замовлення:
Ім'я клієнта: {client_name}
Номер телефону: {phone_number}
Адреса доставки: {delivery_address}
Час приготування замовлення: {preparation_time}
Коментар: {comment if comment else 'Немає'}

Чи сплачено замовлення: {'Так' if is_paid else 'Ні'}
Сума до сплати: {amount_due if not is_paid else 'Сплачено'}
                """
                btn_picked_up = InlineKeyboardButton(
                    "Забрав заказ",
                    callback_data=f"picked_up_order_{order_id}"
                )
                markup = InlineKeyboardMarkup([
                    [btn_picked_up]
                ])

                # Відправляємо повідомлення кур'єру
                try:
                    await bot.send_message(chat_id=courier_telegram_id, text=order_message, reply_markup=markup)
                except Exception as e:
                    print(f"Помилка при надсиланні повідомлення кур'єру: {e}")
                    await message.reply("Не вдалося надіслати повідомлення кур'єру. Перевірте Telegram ID.")
                    return

                # Повідомляємо адміністратора про успішне призначення
                await message.reply(
                    f"Замовлення #{order_id} успішно призначено кур'єру {user.first_name} {user.last_name}.")
            else:
                await message.reply("Деталі замовлення не знайдено.")
        except mysql.connector.Error as err:
            await message.reply(f"Сталася помилка при призначенні кур'єра: {err}")
        finally:
            cursor.close()
            # Скидаємо стан користувача
            user_states.pop(user_id, None)


@bot.on_callback_query(filters.regex(r'^picked_up_order_(\d+)$'))
async def picked_up_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^picked_up_order_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        try:
            # Отримуємо замовлення
            cursor.execute("""
                SELECT courier_id, is_picked_up, is_delivered
                FROM orders
                WHERE id = %s
            """, (order_id,))
            order = cursor.fetchone()
            if not order:
                await callback_query.answer("Замовлення не знайдено.", show_alert=True)
                return

            courier_id_db, is_picked_up, is_delivered = order
            print(order_id)
            # Перевіряємо, чи замовлення призначене цьому кур'єру
            cursor.execute("SELECT courier_id FROM orders WHERE id = %s", (order_id,))
            courier = cursor.fetchone()
            print(courier)
            if not courier or courier[0] != user_id:
                await callback_query.answer("Це замовлення вам не призначене.", show_alert=True)
                return

            if is_picked_up:
                await callback_query.answer("Замовлення вже позначено як забране.", show_alert=True)
                return

            # Оновлюємо статус замовлення
            cursor.execute("""
                UPDATE orders
                SET is_picked_up = TRUE
                WHERE id = %s
            """, (order_id,))
            connection.commit()

            await callback_query.answer("Замовлення позначено як забране.", show_alert=True)
            btn_delivered = InlineKeyboardButton(
                "Доставил заказ",
                callback_data=f"delivered_order_{order_id}"
            )
            markup = InlineKeyboardMarkup([
                [btn_delivered]
            ])
            # Можливо, повідомлення адміністратора або оновлення статусу в чаті
            # Наприклад, відредагувати повідомлення кур'єра, видаливши кнопки
            await callback_query.message.edit_reply_markup(reply_markup=markup)
        except mysql.connector.Error as err:
            print(f"Помилка в picked_up_order_callback: {err}")
            await callback_query.answer("Сталася помилка при оновленні замовлення.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)



@bot.on_callback_query(filters.regex(r'^delivered_order_(\d+)$'))
async def delivered_order_callback(bot, callback_query):
    user_id = callback_query.from_user.id
    match = re.match(r'^delivered_order_(\d+)$', callback_query.data)
    if match:
        order_id = int(match.group(1))
        cursor = connection.cursor()
        try:
            # Отримуємо замовлення
            cursor.execute("""
                SELECT courier_id, is_picked_up, is_delivered
                FROM orders
                WHERE id = %s
            """, (order_id,))
            order = cursor.fetchone()
            if not order:
                await callback_query.answer("Замовлення не знайдено.", show_alert=True)
                return

            courier_id_db, is_picked_up, is_delivered = order

            # Перевіряємо, чи замовлення призначене цьому кур'єру
            cursor.execute("SELECT courier_id FROM orders WHERE id = %s", (order_id,))
            courier = cursor.fetchone()
            print(courier)
            if not courier or courier[0] != user_id:
                await callback_query.answer("Це замовлення вам не призначене.", show_alert=True)
                return

            if not is_picked_up:
                await callback_query.answer("Замовлення ще не позначено як забране.", show_alert=True)
                return

            if is_delivered:
                await callback_query.answer("Замовлення вже позначено як доставлене.", show_alert=True)
                return

            # Оновлюємо статус замовлення
            cursor.execute("""
                UPDATE orders
                SET is_delivered = TRUE
                WHERE id = %s
            """, (order_id,))
            connection.commit()

            await callback_query.answer("Замовлення позначено як доставлене.", show_alert=True)

            # Можливо, повідомлення адміністратора або оновлення статусу в чаті
            # Наприклад, відредагувати повідомлення кур'єра, видаливши кнопки
            await callback_query.message.edit_reply_markup(reply_markup=None)

            # (Опціонально) Повідомити адміністратора або оновити статус замовлення в іншому чаті
            # Наприклад:
            # admin_chat_id = -123456789  # Замініть на ваш чат ID
            # await bot.send_message(admin_chat_id, f"Замовлення #{order_id} було доставлено.")
        except mysql.connector.Error as err:
            print(f"Помилка в delivered_order_callback: {err}")
            await callback_query.answer("Сталася помилка при оновленні замовлення.", show_alert=True)
        finally:
            cursor.close()
    else:
        await callback_query.answer("Невідома команда.", show_alert=True)


async def new_order_alarm(bot, message):
    user_id = message.from_user.id
    order_data = user_states[user_id]['order_data']

    # Automatically assign restaurant_id as the user ID (who sent the message)
    restaurant_id = user_id

    try:
        cursor = connection.cursor()

        # Вставляем новый заказ в таблицу orders и получаем его ID
        cursor.execute("""
            INSERT INTO orders (restaurant_id, client_name, phone_number, delivery_address, preparation_time, comment, 
            is_paid, amount_due)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            restaurant_id,
            order_data['client_name'],
            order_data['phone_number'],
            order_data['delivery_address'],
            order_data['preparation_time'],
            order_data['comment'],
            order_data['is_paid'],
            order_data['amount_due']
        ))
        connection.commit()

        # Получаем ID только что вставленного заказа
        cursor.execute("SELECT LAST_INSERT_ID()")
        order_id = cursor.fetchone()[0]

        # Получаем информацию о ресторане
        sql_restaurant = "SELECT name, address, phone FROM restaurants WHERE id = %s"
        cursor.execute(sql_restaurant, (restaurant_id,))
        restaurant_info = cursor.fetchone()

        if not restaurant_info:
            await message.reply("Не удалось найти информацию о ресторане!")
            return

        restaurant_name, restaurant_address, restaurant_phone = restaurant_info

        # Формируем сообщение для администраторов
        admin_message = f"""
Нове замовлення від ресторану {restaurant_name}
Адреса ресторану: {restaurant_address}
Контактний номер ресторану: {restaurant_phone}

Ім'я клієнта: {order_data['client_name']}
Номер телефону: {order_data['phone_number']}
Адреса доставки: {order_data['delivery_address']}
Час приготування замовлення: {order_data['preparation_time']}
Коментар: {order_data['comment']}

Чи сплачено замовлення: {'Так' if order_data['is_paid'] else 'Ні'}
Сума до сплати: {order_data['amount_due'] if not order_data['is_paid'] else 'Сплачено'}
        """

        # Извлекаем всех администраторов из таблицы admins
        cursor.execute("SELECT telegram_id FROM admins")
        admin_ids = [row[0] for row in cursor.fetchall()]

        btn_edit = InlineKeyboardButton('Редагувати замовлення', callback_data=f"edit_order_{order_id}")
        btn_send = InlineKeyboardButton('Відправити в чат', callback_data=f"send_order_{order_id}")
        btn_send_worker = InlineKeyboardButton('Надіслати курєру', callback_data=f"send_worker_order_{order_id}")
        kb_order_actions = InlineKeyboardMarkup([
            [btn_edit],
            [btn_send],
            [btn_send_worker]
        ])

        # Отправляем сообщение всем администраторам
        for admin_id in admin_ids:
            try:
                await bot.send_message(admin_id, admin_message, reply_markup=kb_order_actions)
                print(f"Повідомлення надіслано адміністратору з ID {admin_id}")
            except Exception as e:
                print(f"Помилка при надсиланні повідомлення адміністратору {admin_id}: {e}")

        cursor.close()

    except mysql.connector.Error as err:
        await message.reply(f"Сталася помилка при збереженні замовлення: {err}")
        return
    finally:
        # Clear the user's state after the order is saved
        user_states.pop(user_id, None)

bot.run()
